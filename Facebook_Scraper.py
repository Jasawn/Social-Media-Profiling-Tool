from time import sleep
from openpyxl.styles import Alignment
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import os.path
import openpyxl
import datetime
from Linkedin_Scraper import initialisedWebScraping


def login_facebook(driver):

    driver.get("http://www.facebook.com")

    username = driver.find_element(By.XPATH, "//input[@type='text']")
    password = driver.find_element(By.XPATH, "//input[@type='password']")

    # enter username and password
    username.clear()
    username.send_keys("@gmail.com")
    password.clear()
    password.send_keys("test")

    # target the login button and click it
    button = driver.find_element(By.XPATH,"//button[@name='login']")
    button.click()

    sleep(2)
    return driver


def facebook_search(driver, fname, lname=None, info=None):

    filteredLinks = []
    fullQuery = "https://www.facebook.com/search/people/?q=" + fname
    if lname:
        fullQuery = fullQuery + "%20" + lname
    if info:
        fullQuery = fullQuery + "%20" + info

    driver.get(fullQuery)
    try:
        allSearch = driver.find_element(By.XPATH, "//div[@class='x78zum5 xdt5ytf x4cne27 xifccgj']")
        print("No results")
        print(allSearch.text)
    except:
        # if need more results, uncomment this line
        # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        allSearch = driver.find_elements(By.XPATH, "//a[@class='x1i10hfl xjbqb8w x6umtig x1b1mbwd xaqea5y xav7gou x9f619 x1ypdohk xt0psk2 xe8uvvx xdj266r x11i5rnm xat24cr x1mh8g0r xexx8yu x4uap5 x18d9i69 xkhd6sd x16tdsg8 x1hl2dhg xggy1nq x1a2a7pz xt0b8zv xzsf02u x1s688f']")
        for links in allSearch:
            unfilteredLink = links.get_attribute('href').split("/")[3:4][0]
            if "?" not in unfilteredLink:
                filteredLinks.append(unfilteredLink)

    return filteredLinks


def google_facebook(fname, lname, info=None):
    filteredLink = []
    driver = initialisedWebScraping()
    driver.get("https://www.google.com")
    filters = "https://www.facebook.com"

    fullName = fname + " " + lname
    searchQuery = "site:facebook.com AND " + fullName
    if info:
        searchQuery = searchQuery + " AND " + info

    searchBar = driver.find_element(By.NAME, 'q')
    searchBar.send_keys(searchQuery)
    sleep(1)
    searchBar.send_keys(Keys.RETURN)
    try:
        print(driver.find_element(By.XPATH, '//div[@id="topstuff"]/div/div/p').text)
        result = "No results returned"
    except:
        try:
            allLinks = driver.find_elements(By.TAG_NAME, "a")
            for links in allLinks:
                unfilteredLink = str(links.get_attribute('href'))
                if unfilteredLink.startswith(filters):
                    unfilteredLink = unfilteredLink.split("/")
                    if len(unfilteredLink) == 4:
                        filteredLink.append(unfilteredLink)
        except:
            print("Error")

    print(unfilteredLink)

def profile_facebook(driver, profileID):

    link = "https://www.facebook.com/" + profileID
    # for work
    companyList = []
    workPositionList = []
    workDurationList = []
    workLocationList = []

    # for university
    universityNameList = []
    universityDegreeList = []
    universityDurationList = []

    # for high school
    highSchoolNameList = []
    highSchoolDurationList = []

    # for location
    locationList = []
    locationTimeList = []

    # scraping initial information about profile
    driver.get(link)
    profilePictureLink = driver.find_elements(By.TAG_NAME, "image")[1].get_attribute('xlink:href')
    profileName = driver.find_element(By.XPATH, "//h1[@class='x1heor9g x1qlqyl8 x1pd3egz x1a2a7pz']").text

    # Work and Education page
    driver.get(link + "/about_work_and_education")

    # retrieving information from Work section of profile
    workSection = driver.find_element(By.XPATH, '//div[@class="x1gan7if"][1]')
    # selecting individual work entries
    workList = workSection.find_elements(By.XPATH, './/div[@class="x13faqbe"]')
    # Finding companies listed on profile
    for workIndiv in workList:
        workIndiv = workIndiv.text.split("\n")
        # print(workIndiv)
        workDuration = "No duration available"
        workLocation = "No work location available"
        company = "No company to show"
        workPosition = "No work position available"
        try:
            if workIndiv[0] != "No workplaces to show":
                if workIndiv[0].split(" at ", 1)[0] != "Works" and workIndiv[0].split(" at ", 1)[0] != "Worked":
                    workPosition = workIndiv[0].split(" at ", 1)[0]
                company = workIndiv[0].split("at ", 1)[1]

                try:
                    if workIndiv[1]:
                        if " - " in workIndiv[1]:
                            workDuration = workIndiv[1]
                        else:
                            workLocation = workIndiv[1]
                except:
                    pass

                try:
                    if (workIndiv[3]):
                        workLocation = workIndiv[3]
                except:
                    pass
        except:
            company = workIndiv[0]
        companyList.append(company + "\n")
        workPositionList.append(workPosition + "\n")
        workDurationList.append(workDuration + "\n")
        workLocationList.append(workLocation + "\n")

    # retrieving information from University section of profile
    universitySection = driver.find_element(By.XPATH, '//div[@class="x1gan7if"][2]')
    # selecting individual university entries
    universityAll = universitySection.find_elements(By.XPATH, './/div[@class="x13faqbe"]')
    # Finding university/school listed on profile
    for universityIndiv in universityAll:
        universityName = "No universities listed"
        universityDegree = "No university degree to show"
        universityDuration = "No university duration to show"
        universityIndiv = universityIndiv.text.split("\n")

        for i in universityIndiv:
            if i == "·":
                universityIndiv.remove(i)

        universityIndivCleaned = [i for i in universityIndiv if "Also studied " not in i]

        if universityIndivCleaned[0] != "No schools/universities to show":
            universityName = universityIndiv[0].split("at ", 1)[1]

            try:
                if universityIndivCleaned[1]:
                    if " - " in universityIndiv[1] or "year" in universityIndiv[1]:
                        universityDuration = universityIndiv[1]
                    else:
                        universityDegree = universityIndiv[1]
            except:
                pass

            try:
                if universityIndivCleaned[2]:
                    universityDuration = universityIndiv[2]
            except:
                pass

        universityNameList.append(universityName + "\n")
        universityDegreeList.append(universityDegree + "\n")
        universityDurationList.append(universityDuration + "\n")

    # retrieving information from High School section of profile
    highSchoolSection = driver.find_elements(By.XPATH, '//div[@class="x1gan7if"][2]/following-sibling::div')
    for highSchoolExp in highSchoolSection:
        # selecting individual high school entries
        highSchoolAll = highSchoolExp.find_elements(By.XPATH, './/div[@class="x13faqbe"]')
        # Finding high schools listed on profile
        for highSchoolIndiv in highSchoolAll:
            highSchoolIndiv = highSchoolIndiv.text.split("\n")
            highSchoolDuration = "No high school duration available"
            highSchoolName = "No high school available"
            if highSchoolIndiv[0] != "No schools/universities to show":
                highSchoolName = highSchoolIndiv[0].split("to ", 1)[1]
                try:
                    if highSchoolIndiv[1]:
                        highSchoolDuration = highSchoolIndiv[1]
                except:
                    pass

            highSchoolNameList.append(highSchoolName + "\n")
            highSchoolDurationList.append(highSchoolDuration + "\n")

    # retrieving information about locations
    driver.get(link + "/about_places")
    locationAndTime = driver.find_elements(By.XPATH, "//div[@class='x13faqbe']")
    currentLocation = "No current location available"
    homeTown = "No current home town available"
    for locationAndTimeIndiv in locationAndTime:
        if locationAndTimeIndiv.text != "No places to show":
            locationAndTimeIndiv = locationAndTimeIndiv.text.split("\n")
            try:
                if locationAndTimeIndiv[1] == "Current town/city":
                    # print("matching current city")
                    currentLocation = locationAndTimeIndiv[0]
                    # print(currentLocation)
                if locationAndTimeIndiv[1] == "Home town":
                    homeTown = locationAndTimeIndiv[0]
            except:
                pass

    # retrieving information about contact
    driver.get(link + "/about_contact_and_basic_info")

    residentialAddress = "Address not available"
    phoneNumber = "Phone number not available"
    emailAddress = "No email available"

    fullContactInfo = driver.find_element(By.XPATH, "//div[@class='x1gan7if'][1]").text.split("\n")[1:]
    for i in range(len(fullContactInfo)):
        if fullContactInfo[i] == "Mobile":
            phoneNumber = fullContactInfo[i-1]
        if fullContactInfo[i] == "Address":
            residentialAddress = fullContactInfo[i-1]
        if fullContactInfo[i] == "Email":
            emailAddress = fullContactInfo[i-1]

    # retrieving all websites or links available on profile
    websiteAvailable = "No website available"
    linkedInstagramAccount = "No instagram account linked"
    linkedTwitterAccount = "No twitter account linked"
    linkedLinkedInAccount = "No linkedIn account linked"

    linksAvailable = driver.find_element(By.XPATH, "//div[@class='x1gan7if'][2]").text.split("\n")[1:]

    for i in range(len(linksAvailable)):
        if linksAvailable[i] == "Website":
            websiteAvailable = linksAvailable[i-1]
        if linksAvailable[i] == "Instagram":
            linkedInstagramAccount = linksAvailable[i-1]
        if linksAvailable[i] == "Twitter":
            linkedTwitterAccount = linksAvailable[i-1]
        if linksAvailable[i] == "LinkedIn":
            linkedLinkedInAccount = linksAvailable[i-1]

    # finding basic information
    birthDateAvailable = "Birth Date not available"
    birthYearAvailable = "Birth Year not available"
    languagesAvailable = "Languages not available"
    religionAvailable = "Religion not available"
    politicsAvailable = "Politics not available"
    genderAvailable = "Gender not available"

    basicInfoAvailable = driver.find_element(By.XPATH, "//div[@class='x1gan7if'][2]/following-sibling::div").text.split("\n")
    for i in range(len(basicInfoAvailable)):
        if basicInfoAvailable[i] == "Gender":
            genderAvailable = basicInfoAvailable[i-1]
        if basicInfoAvailable[i] == "Birth date":
            birthDateAvailable = basicInfoAvailable[i-1]
        if basicInfoAvailable[i] == "Birth year":
            birthYearAvailable = basicInfoAvailable[i-1]
        if basicInfoAvailable[i] == "Languages":
            languagesAvailable = basicInfoAvailable[i-1]
        if basicInfoAvailable[i] == "Religious views":
            religionAvailable = basicInfoAvailable[i-1]
        if basicInfoAvailable[i] == "Political views":
            politicsAvailable = basicInfoAvailable[i-1]
    driver.get(link + "/about_details")
    bioInfoAvailable = driver.find_element(By.XPATH, "//div[@class='x1gan7if'][1]").text.split("\n")
    bioAvailable = "No bio available"
    if bioInfoAvailable[1] != "No additional details to show":
        bioAvailable = bioInfoAvailable[1]

    facebook_data_dict = {
        "profileID": link,
        "profileName": profileName,
        "profilePictureLink": profilePictureLink,
        "bioAvailable": bioAvailable,
        "genderAvailable": genderAvailable,
        "companyList": companyList,
        "workPositionList": workPositionList,
        "workDurationList": workDurationList,
        "workLocationList": workLocationList,
        "universityNameList": universityNameList,
        "universityDegreeList": universityDegreeList,
        "universityDurationList": universityDurationList,
        "highSchoolNameList": highSchoolNameList,
        "highSchoolDurationList": highSchoolDurationList,
        "currentLocation": currentLocation,
        "homeTown": homeTown,
        "residentialAddress": residentialAddress,
        "phoneNumber": phoneNumber,
        "emailAddress": emailAddress,
        "websiteAvailable": websiteAvailable,
        "linkedInstagramAccount": linkedInstagramAccount,
        "linkedTwitterAccount": linkedTwitterAccount,
        "linkedLinkedInAccount": linkedLinkedInAccount,
        "birthDateAvailable": birthDateAvailable,
        "birthYearAvailable": birthYearAvailable,
        "languagesAvailable": languagesAvailable,
        "religionAvailable": religionAvailable,
        "politicsAvailable": politicsAvailable
    }
    return facebook_data_dict

def writeToExcel(dataDict, searchName):
    filename = str(datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")).replace("-", "") + " " + searchName + "_FacebookData.xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        workbook.save(filename)

    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.worksheets[0]
    column = 1
    header = ["Profile ID", "Name", "Profile Picture Link", "Bio", "Gender", "Companies", "Work Position", "Work Duration", "Work Location", "Universities", "University Degree", "University Duration", "High Schools", "High School Duration", "Current Location", "Home Town", "Residential Address", "Phone Number", "Email Address", "Websites", "Instagram", "Twitter", "LinkedIn", "Birth Date", "Birth Year", "Languages", "Religion", "Politics"]

    for headerData in header:
        worksheet.cell(1, column).value = headerData
        column += 1
    row = 2

    for id, info in dataDict.items():
        column = 1
        for key, value in info.items():
            if type(value) is list:
                worksheet.cell(row, column).alignment = Alignment(wrapText=True)
                toBeInserted = ""
                for i in range(len(value)):
                    toBeInserted += value[i]
                worksheet.cell(row, column).value = toBeInserted
            else:
                worksheet.cell(row, column).value = value
            column += 1
        row += 1

    workbook.save(filename)
    workbook.close()

    return filename

def facebookSearchAndScrape(fname, lname, info=None):
    driver = login_facebook(initialisedWebScraping())
    searchName = fname + " " + lname
    profilesToScrape = facebook_search(driver, fname, lname, info)
    facebookDataDict = {}
    for profile in profilesToScrape:
        facebookDataDict.update({profile: profile_facebook(driver, profile)})
    return writeToExcel(facebookDataDict, searchName)

