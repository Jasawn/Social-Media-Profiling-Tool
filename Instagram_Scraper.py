from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment
from time import sleep
from Linkedin_Scraper import *
import re
import os.path
import openpyxl
import datetime

def GoogleSearch_Instagram(fname,lname,username,country):
    ## Filter for google dorking
    filters = re.compile("www.instagram.com")
    filteredLink = []

    searchName = fname + " " + lname
    dorking = f'site:instagram.com AND {searchName}'
    if country:
        dorking = f'site:instagram.com AND {searchName} AND {country}'

    ## Start Selenium
    driver = initialisedWebScraping()

    ## Navigate to the page specified below
    driver.get('https://www.google.com')

    enterSearchQuery = driver.find_element(By.NAME,'q')
    enterSearchQuery.send_keys(dorking)
    sleep(1)
    enterSearchQuery.send_keys(Keys.RETURN)
    sleep(3)

    try:
        driver.find_element(By.XPATH, '//div[@id="topstuff"]/div/div/p').text
        return "Error"
    except:
        allLinks = driver.find_elements(By.TAG_NAME,"a")
        for links in allLinks:
            unfilteredLink = links.get_attribute('href')
            if unfilteredLink and re.match(filters,unfilteredLink.split('/')[2]) and not unfilteredLink.split('/')[3] == 'p':
                filteredLink.append(unfilteredLink)
        if username:
            return(instagramLogin(driver,filteredLink,username,searchName))
        else:
            return(instagramLogin(driver,filteredLink,'',searchName))

def GoogleNextPage(driver):
    nextPageButton = driver.find_element(By.ID,'pnnext')
    nextPageButton.click()

def instagramLogin(driver, filteredLink,username,searchName):
    ## Navigate to the page specified below
    driver.get('https://www.instagram.com/')
    driver.implicitly_wait(3)

    ## To simulate key strokes and enter the email details
    enterEmail = driver.find_element(By.NAME,'username')
    enterEmail.send_keys('@live.com')
    sleep(3)

    ## To simulate key strokes and enter the password
    enterPassword = driver.find_element(By.NAME,'password')
    enterPassword.send_keys('test')
    sleep(3)

    ## Locate submit button using XPath
    log_in_button = driver.find_element(By.XPATH,'//*[@type="submit"]')
    log_in_button.click()
    sleep(8)

    try:
        notNow = driver.find_element(By.XPATH,'//main[@role="main"]/div/div/div/div/button')
        notNow.click()
    except:
        return "Error"

    return(InstagramSearchScraping(driver,filteredLink,username,searchName))

def InstagramSearchScraping(driver,filteredLink,username,searchName):
    searchBarButton = driver.find_element(By.XPATH,"//*[local-name() = 'svg' and @aria-label='Search']")
    searchBarButton.click()
    searchBar = driver.find_element(By.XPATH,'//input[@aria-label="Search input"]')
    searchBar.send_keys(username)
    sleep(2)

    profileCount = 0
    profileLinks = []

    scrapedInfo = {}

    try:
        searchProfileLinks = driver.find_elements(By.XPATH,'//input[@aria-label="Search input"]/../../../div[2]/div/div/div/a')
        for links in searchProfileLinks:
            if profileCount == 5:
                break
            else:
                profileCount += 1
            if (links.get_attribute('href') not in filteredLink):
                profileLinks.append(links.get_attribute('href'))

        for link in profileLinks:
            driver.get(link)
            sleep(3)
            profileInfo = []
            try:
                isPrivate = driver.find_element(By.XPATH,'//main[@role="main"]/div/div/article/div/div/h2').text
                profilePic = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/div/div/div/button/img')
                profilePicLink = profilePic.get_attribute('src')

                try:
                    name = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/span').text
                except:
                    name = "None"

                try:
                    job = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div').text
                    bio = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div[2]').text
                except:
                    job = "None"
                    try:
                        bio = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div').text
                    except:
                        bio = "None"

                try:
                    additionalLink = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/a')
                    addLink = additionalLink.get_attribute('href')
                except:
                    addLink = "None"

                photoSrcLink = ["None","None"]
                profileInfo.append(name)
                profileInfo.append(profilePicLink)
                profileInfo.append(job)
                profileInfo.append(bio)
                profileInfo.append(addLink)
                profileInfo.append(photoSrcLink)

                scrapedInfo.update({link:profileInfo})
            except:
                filteredLink.append(link)
    except:
        profileLink = "No results found!"

    return(instagramScraping(driver,filteredLink,scrapedInfo,searchName))

def instagramScraping(driver,filteredLink,scrapedInfo,searchName):
    for profileLink in filteredLink:
        profileInfo = []
        photoSrcLink = []
        ## Redirect to the target's profile
        driver.get(profileLink)
        sleep(4)

        ## Retrieve Profile Image
        try:
            profilePic = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/div/div/span/img')
            profilePicLink = profilePic.get_attribute('src')
        except:
            profilePic = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/div/div/div/button/img')
            profilePicLink = profilePic.get_attribute('src')
        ## Retrieve Biography / Job
        try:
            job = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div').text
            bio = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div[2]').text
        except:
            job = "None"
            try:
                bio = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/div').text
            except:
                bio = "None"
        ## Retrieve Name
        try:
            name = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/span').text
        except:
            name = "None"
        ## Additional Link
        try:
            additionalLink = driver.find_element(By.XPATH,'//main[@role="main"]/div/header/section/div[3]/a')
            link = additionalLink.get_attribute('href')
        except:
            link = "None"

        profileInfo.append(name)
        profileInfo.append(profilePicLink)
        profileInfo.append(job)
        profileInfo.append(bio)
        profileInfo.append(link)

        ## Retrieve first few recent posts of the profile
        try:
            photos = driver.find_elements(By.XPATH,'//main[@role="main"]/div/div[2]/article/div[1]/div/div[1]/div/a')
            if len(photos) == 0:
                try:
                    ## Different format
                    photos = driver.find_elements(By.XPATH,'//main[@role="main"]/div/div[3]/article/div/div/div[1]/div/a')
                    if len(photos) == 0:
                        photoSrcLink = ["None","None"]
                    else:
                        photoSrcLink = ImgXPATH(driver,photos,"Alt")
                except:
                    print("In except1")
            else:
                photoSrcLink = ImgXPATH(driver,photos,"Main")
        except:
            print("Except1")
            photoSrcLink = ["None","None"]

        try:
            photos = driver.find_elements(By.XPATH,'//main[@role="main"]/div/div[2]/article/div[1]/div/div[2]/div/a')
            if len(photos) == 0:
                try:
                    ## Different format
                    photos = driver.find_elements(By.XPATH,'//main[@role="main"]/div/div[3]/article/div/div/div[2]/div/a')
                    if len(photos) == 0:
                        if photoSrcLink == []:
                            photoSrcLink = ["None","None"]
                    else:
                        photoSrcLink = photoSrcLink + ImgXPATH(driver,photos,"Alt")
                except:
                    print("In except2")
            else:
                photoSrcLink = photoSrcLink + ImgXPATH(driver,photos,"Main")
        except:
            photoSrcLink = photoSrcLink
        profileInfo.append(photoSrcLink)
        scrapedInfo.update({profileLink:profileInfo})

    # logout function
    profileButton = driver.find_element(By.XPATH, "//*[local-name() = 'svg' and @aria-label='Settings' and @role='img']/..")
    profileButton.click()
    logout = driver.find_element(By.XPATH, "//*[local-name() = 'svg' and @aria-label='Settings' and @role='img']/../../../../../../div[1]/div/div/div[5]")
    logout.click()
    driver.quit()

    if scrapedInfo:
        return(writeToExcel(scrapedInfo,searchName))
    else:
        return("Error")

def ImgXPATH(driver, photos, case):
    xpath = []
    if case == "Main":
        multipleImageSrc = '//article["presentation"]/div/div[1]/div/div[1]/div[2]/div/div/div/ul/li/div/div/div/div/div[1]/img'
        multipleImageSrcAlt = ''
    else:
        multipleImageSrc = '//article["presentation"]/div/div[1]/div/div[1]/div[2]/div/div/div/ul/li/div/div/div/div/div[1]/div[1]/img'
        multipleImageSrcAlt = '//article["presentation"]/div/div[1]/div/div[1]/div[2]/div/div/div/ul/li/div/div/div/div/div[1]/img'

    singleImg = '//article["presentation"]/div/div[1]/div/div/div/div[1]/img'
    singleImgLoc = '//article["presentation"]/div/div[2]/div/div/div[1]/div/header/div[2]/div[2]/div[2]/div/a'
    numberofMultipleImage = '//article["presentation"]/div/div[1]/div/div[2]/div'
    multipleImageSrcLoc = '//article["presentation"]/div/div[2]/div/div/div[1]/div/header/div[2]/div[2]/div[2]/div/a'
    button = '//article["presentation"]/div/div[1]/div/div[1]/div[2]/div/button[@aria-label="Next"]'

    xpath.append(singleImg)
    xpath.append(singleImgLoc)
    xpath.append(numberofMultipleImage)
    xpath.append(multipleImageSrc)
    xpath.append(multipleImageSrcLoc)
    xpath.append(button)
    xpath.append(multipleImageSrcAlt)

    return(instagramScrapingRecentPhotos(driver,photos,xpath))

def instagramScrapingRecentPhotos(driver,photos,xpath):
    photoSrcLink = []
    for photo in photos:
        photo.click()
        sleep(3)

        ## Retrieve single image
        try:
            ## Retrieve single image URL
            singleImg = photo.find_element(By.XPATH, xpath[0])
            imgUrl = singleImg.get_attribute('src')

            ## Retrieve single image location
            try:
                singleImgLoc = photo.find_element(By.XPATH, xpath[1]).text
            except:
                singleImgLoc = "None"
            photoSrcLink.append(imgUrl)
            photoSrcLink.append(singleImgLoc)
        except:
            ## Retrieve multiple image
            try:
                numberOfMultipleImage =  photo.find_elements(By.XPATH, xpath[2])
                button = photo.find_element(By.XPATH, xpath[5])
                multipleImgUrl = []
                multipleImgLoc = []
                for loop in numberOfMultipleImage:
                    multipleImageSrc = photo.find_elements(By.XPATH, xpath[3])
                    if len(multipleImageSrc) == 0:
                        multipleImageSrc = photo.find_elements(By.XPATH, xpath[6])
                    for imageUrl in multipleImageSrc:
                        if imageUrl.get_attribute('src') not in multipleImgUrl:
                            multipleImgUrl.append(imageUrl.get_attribute('src'))
                        else:
                            continue
                    try:
                        multipleImageSrcLoc = photo.find_element(By.XPATH, xpath[4]).text
                        multipleImgLoc.append(multipleImageSrcLoc)
                    except:
                        multipleImgLoc.append("None")
                    try:
                        button.click()
                    except:
                        print("End")
                    sleep(3)

                itemCount = 0
                for items in multipleImgUrl:
                    photoSrcLink.append(items)
                    photoSrcLink.append(multipleImgLoc[itemCount])
                    itemCount += 1
            except:
                photoSrcLink.append("None")
                photoSrcLink.append("None")
        closeButton = photo.find_element(By.XPATH,"//*[local-name() = 'svg' and @aria-label='Close']")
        closeButton.click()
        sleep(3)

    return photoSrcLink

def writeToExcel(scrapedInfo, searchName):
    filename = str(datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")).replace("-","") + " " + searchName + " InstaData.xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        workbook.save(filename)

    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.worksheets[0]
    header = ["Profile Link", "Name", "Profile Pic Link","Job","Biography","Additional Link", "Recent Photos"]

    column = 1

    ## Write Header content to excel
    for headerData in header:
        worksheet.cell(1,column).value = headerData
        column += 1
    row = 2

    ## Write the dictionary contents into the excel
    for key,allData in scrapedInfo.items():
        column = 1
        worksheet.cell(row,column).value = key
        column += 1

        for data in allData:
            if column == 7:
                while data:
                    if isinstance(data[0],list):
                        while data[0]:
                            try:
                                value = worksheet.cell(row,column).value
                                newValue = value + "\n" + data[0][0] + "\n" + data[0][1] + "\n"
                            except:
                                newValue = data[0][0] + "\n" + data[0][1] + "\n"
                            worksheet.cell(row,column).value = newValue
                            data[0].pop(0)
                            data[0].pop(0)
                    else:
                        try:
                            value = worksheet.cell(row,column).value
                            newValue = value + "\n" + data[0] + "\n" + data[1] + "\n"
                        except:
                            newValue = data[0] + "\n" + data[1] + "\n"
                        worksheet.cell(row,column).value = newValue

                    data.pop(0)
                    data.pop(0)

                if data == []:
                    worksheet.cell(row,column).alignment = Alignment(wrapText=True)
            else:
                worksheet.cell(row,column).value = data
            column += 1
        row += 1
    workbook.save(filename)
    workbook.close()
    return filename
