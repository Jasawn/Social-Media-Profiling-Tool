from flask import Flask, redirect, url_for, render_template, request, jsonify, flash
from sentence_transformers import SentenceTransformer, util
from Linkedin_Scraper import *
from Instagram_Scraper import *
from Twitter_Scraper import *
from Facebook_Scraper import *
import os
import openpyxl
import requests
import shutil
import glob
import torch
from IPython.display import clear_output
from PIL import Image

godEye = Flask(__name__)
imageFolder = 'image'

godEye.config["IMAGE_UPLOADS"] = imageFolder
godEye.secret_key = "SECRETTTTT"

@godEye.route("/")
def landingPage():
    information = {}
    return render_template("index.html")                           # Main page for the user to enter the required information for OSINT

@godEye.route("/home")
def homePage():
    return render_template("searchpage.html")

@godEye.route("/searchResult", methods=['POST'])
def searchResult():
    data = {}

    if os.path.exists(imageFolder):
        shutil.rmtree(imageFolder)

    os.mkdir(imageFolder)

    if request.files['image']:
        uploadedImage = request.files['image']
        uploadedImage.save(os.path.join(godEye.config["IMAGE_UPLOADS"],uploadedImage.filename))         # Save the file user have uploaded
        print ("Image file is located at: " + godEye.config["IMAGE_UPLOADS"],uploadedImage.filename)
        crop("image/"+uploadedImage.filename, "0")

    ## All the variables user have entered
    if request.form['username'] or request.form['fname'] or request.form['lname'] or request.form['country'] or request.form['edu'] or request.form['relevantInfo'] or request.form['country']:
        information = {'Username' : request.form['username'] , 'First name' : request.form['fname'] , 'Last name':request.form['lname'],
        'relevantInfo' : request.form['relevantInfo'] , 'Education/Organisation': request.form['edu'] , 'Country' : request.form['country']}

        ## Here to run scraping codes and object matching
        twitterFilename = twitterLogin(information['First name'], information['Last name'])
        linkedInFilename = GoogleSearch_LinkedIn(information['First name'],information['Last name'],information['Education/Organisation'])
        instagramFilename = GoogleSearch_Instagram(information['First name'],information['Last name'],information['Username'],information['Country'])
        facebookFilename = facebookSearchAndScrape(information['First name'], information['Last name'], information['relevantInfo'])

        if facebookFilename == "Error":
            data = {'Facebook': []}
        else:
            try:
                data = readFacebook(data, facebookFilename)
            except:
                data = {'Facebook': []}

        if linkedInFilename == "Error":
           data = {'LinkedIn' : []}
        else:
           try:
               data = readLinkedIn(data,linkedInFilename)
           except:
               data = {'LinkedIn' : []}

        if instagramFilename == "Error":
            data = {'Instagram': []}
        else:
            try:
                data = readInstagram(data,instagramFilename)
            except:
                data = {'Instagram': []}

        if twitterFilename == "Error":
           data = {'Twitter' : []}
        else:
            try:
                data = readTwitter(data,twitterFilename)
            except:
                data = {'Twitter' : []}

        searchResult = information['First name'] + information['Last name']
        data.update({'Name':searchResult})

        fileName = twitterFilename + "!!!***" + linkedInFilename + "!!!***" + instagramFilename + "!!!***" + facebookFilename
        data.update({'Filename':fileName})

        path = "reference_image"
        shutil.rmtree(path, ignore_errors=True)
        return render_template("displaypage.html",data=data)
    else:
        return render_template("searchpage.html")

@godEye.route("/saveResult", methods=["POST","GET"])
def saveResult():
    retrieveTime = request.form['Filename'].split("!!!***")[0][0:15]
    destinationPath = f"SavedResults/Result for {request.form['name']} at {retrieveTime}"
    imageDir = "static/image/"

    if not os.path.exists("SavedResults"):
        os.mkdir("SavedResults")

    if os.path.exists(destinationPath):
        return "", 401
    else:
        os.mkdir(destinationPath)
        for filename in request.form['Filename'].split("!!!***"):
            shutil.copy(filename, destinationPath+"/"+filename)
        shutil.copytree(imageDir, destinationPath+"/image")
        return "", 200

@godEye.route("/deleteData", methods=["POST","GET"])
def deleteData():
    for filename in request.form['Filename'].split("!!!***"):
        os.remove(filename)
    shutil.rmtree("static/image/")
    return "", 200

@godEye.route("/successPage", methods=["POST","GET"])
def successPage():
    flash("Data successfully saved!")
    return render_template("saveSuccessPage.html")

def readFacebook(data, filename):
    facebookData = {}
    destinationPath = "static/image/facebook/"

    if not os.path.exists(destinationPath):
        os.makedirs(destinationPath)
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        row = worksheet.max_row
        column = worksheet.max_column
        for i in range(2, row+1):
            datavalue = []
            for j in range(2, column+2):
                content = worksheet.cell(row = i, column=j).value
                ## This is image file
                if j == 3 and content != "None":
                    retreveImg = requests.get(content)
                    identifier = (worksheet.cell(row=i,column=1).value).split('/')[-1]
                    fileName = destinationPath + identifier + ".jpg"
                    imgFile = open(fileName, "wb")
                    imgFile.write(retreveImg.content)
                    imgFile.close
                    htmlFileName = identifier + ".jpg"
                    ## Give the location of the image file
                    content = htmlFileName
                    path = "image"
                    dir = os.listdir(path)
                    if len(dir) != 0:
                        crop("static/image/facebook/"+content, "4")
                        percentage = compare("4")
                        worksheet.cell(row=1, column=column + 1).value = "Percentage Match"
                        worksheet.cell(row=i, column=column + 1).value = str(percentage)
                        workbook.save(filename)
                if content:
                    content = content.replace("\n","<br>")
                datavalue.append(content)
                print(worksheet.cell(row=2, column=51).value)
                print("-------------------------")
            facebookData.update({worksheet.cell(row=i,column=1).value:datavalue})
        workbook.close()
    except:
        print("error")
        facebookData = {}
    data.update({"Facebook": facebookData})
    return data


def readInstagram(data,filename):
    instagramData = {}
    destinationPath = "static/image/instagram/"

    if not os.path.exists(destinationPath):
        os.makedirs(destinationPath)
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        row = worksheet.max_row
        column = worksheet.max_column
        for i in range(2, row+1):
            datavalue = []
            for j in range(2, column+2):
                content = worksheet.cell(row = i, column=j).value
                multipleImage = []
                imageCounter = 0
                if j == 3 and content != "None":
                    retrieveImg = requests.get(content)
                    identifier = (worksheet.cell(row=i,column=1).value).split('/')[-2]
                    fileName = "static/image/instagram/" + identifier + ".jpg"
                    imgFile = open(fileName, "wb")
                    imgFile.write(retrieveImg.content)
                    imgFile.close
                    htmlFileName = identifier + ".jpg"
                    ## Give the location of the image file
                    content = htmlFileName
                    path = "image"
                    dir = os.listdir(path)
                    if len(dir) != 0:
                        crop("static/image/instagram/"+content, "1")
                        percentage = compare("1")
                        worksheet.cell(row=1, column=column + 1).value = "Percentage Match"
                        worksheet.cell(row=i, column=column + 1).value = str(percentage)
                        workbook.save(filename)
                elif j == 7:
                    images = content.split("\n")
                    while images:
                        if images[0] != "None" and images[0]:
                            retrieveImg = requests.get(images[0])
                            identifierMultiple = identifier + str(imageCounter)
                            fileName = destinationPath + identifierMultiple + ".jpg"
                            imgFile = open(fileName, "wb")
                            imgFile.write(retrieveImg.content)
                            imgFile.close
                            htmlFileName = identifierMultiple + ".jpg"
                            multipleImage.append(htmlFileName)
                            multipleImage.append(images[1].replace("\n","<br>"))
                        images.pop(0)
                        images.pop(0)
                        images.pop(0)
                        imageCounter += 1
                if content and j != 7:
                    content = content.replace("\n","<br>")
                    datavalue.append(content)
                if multipleImage:
                    datavalue.append(multipleImage)
                elif not multipleImage and j == 7:
                    datavalue.append("None <br>")
            print(datavalue)
            instagramData.update({worksheet.cell(row=i,column=1).value:datavalue})
    except:
        instagramData = {}
    data.update({"Instagram":instagramData})
    return data

def readLinkedIn(data,filename):
    linkedInData = {}
    destinationPath = "static/image/linkedin/"

    if not os.path.exists(destinationPath):
        os.makedirs(destinationPath)
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        row = worksheet.max_row
        column = worksheet.max_column
        #16
        print(column)
        for i in range(2, row+1):
            datavalue = []
            for j in range(2, column+2):
                print(j)
                content = worksheet.cell(row = i, column=j).value
                ## This is image file
                if j == 3 and content != "None":
                    retreveImg = requests.get(content)
                    identifier = (worksheet.cell(row=i,column=1).value).split('/')[-1]
                    fileName = destinationPath + identifier + ".jpg"
                    imgFile = open(fileName, "wb")
                    imgFile.write(retreveImg.content)
                    imgFile.close
                    htmlFileName = identifier + ".jpg"
                        ## Give the location of the image file
                    content = htmlFileName
                    path = "image"
                    dir = os.listdir(path)
                    if len(dir) != 0:
                        crop("static/image/linkedin/"+content,"2")
                        percentage = compare("2")
                        worksheet.cell(row=1, column = column+1).value = "Percentage Match"
                        worksheet.cell(row=i, column = column+1).value = str(percentage)
                        workbook.save(filename)
                if content:
                    content = content.replace("\n","<br>")
                datavalue.append(content)
                print(worksheet.cell(row=2, column=17).value)
                print("-------------------------")
            linkedInData.update({worksheet.cell(row=i,column=1).value:datavalue})
        workbook.close()
    except:
        print("error")
        linkedInData = {}
    data.update({"LinkedIn":linkedInData})
    return data


def readTwitter(data,filename):
    twitterData = {}
    destinationPath = "static/image/twitter/"

    if not os.path.exists(destinationPath):
        os.makedirs(destinationPath)
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        row = worksheet.max_row
        column = worksheet.max_column
        for i in range(2, row+1):
            datavalue = []
            for j in range(2, column+2):
                content = worksheet.cell(row = i, column=j).value
                if j == 4 and content:
                    retreveImg = requests.get(content)
                    identifier = (worksheet.cell(row=i,column=9).value).split('/')[-1]
                    fileName = "static/image/twitter/" + identifier + ".jpg"
                    imgFile = open(fileName, "wb")
                    imgFile.write(retreveImg.content)
                    imgFile.close
                    htmlFileName = identifier + ".jpg"
                        ## Give the location of the image file
                    content = htmlFileName
                    path = "image"
                    dir = os.listdir(path)
                    if len(dir) != 0:
                        crop("static/image/twitter/"+content, "3")
                        percentage = compare("3")
                        worksheet.cell(row=1, column=column + 1).value = "Percentage Match"
                        worksheet.cell(row=i, column=column + 1).value = str(percentage)
                        workbook.save(filename)
                if content:
                    content = content.replace("\n","<br>")
                datavalue.append(content)
            twitterData.update({worksheet.cell(row=i, column=1).value:datavalue})
        workbook.close()
    except:
        twitterData = {}
    data.update({"Twitter":twitterData})
    return data

def crop(img, x):
    dict = {
        "0": "reference_image/",
        "1": "static/crop/instagram/",
        "2": "static/crop/linkedin/",
        "3": "static/crop/twitter/",
        "4": "static/crop/facebook/"
    }
    dict1 = {
        "0": "reference_image/crops/head/",
        "1": "static/crop/instagram/crops/head/",
        "2": "static/crop/linkedin/crops/head/",
        "3": "static/crop/twitter/crops/head/",
        "4": "static/crop/facebook/crops/head/"
    }
    # Use custom model to extract face
    model = torch.hub.load('ultralytics/yolov5', 'custom', path='crowdhuman_yolov5m.pt',
                           force_reload=True)
    print(img)
    #imgs = [img]
    # Model's confidence level
    model.conf = 0.8
    # Specify [1] to only extract face
    model.classes = [1]
    # Inference, run the model on the 2 foreground images
    results = model(img, size=320)
    clear_output()
    # Count number of objects
    try:
        val = len(results.pandas().xyxy[0])
        if (val > 0):
            # Save results as new image
            results.crop(save_dir=dict[x])
        else:
            results.crop(save_dir=dict1[x])
    except:
        results.crop(save_dir=dict1[x])
    return 0


def compare(x):
    print('Loading CLIP Model...')
    model = SentenceTransformer('clip-ViT-B-32')  # 86.09
    #model = SentenceTransformer('clip-ViT-L-14')  # 91.63

    dict = {
        "1": "static/crop/instagram/crops/head/*.jpg",
        "2": "static/crop/linkedin/crops/head/*.jpg",
        "3": "static/crop/twitter/crops/head/*.jpg",
        "4": "static/crop/facebook/crops/head/*.jpg"
    }

    list_of_images = []
    #image_1 = Image.open('reference_image/crops/head/' + uploadedimage)
    for filename in glob.glob("reference_image/crops/head/*.jpg"):
        print(filename)
        image_1 = Image.open(filename)
        list_of_images.append(image_1)
    for filename1 in glob.glob(dict[x]):
        image_2 = Image.open(filename1)
        list_of_images.append(image_2)
    encoded_image = model.encode(list_of_images, batch_size=128, convert_to_tensor=True, show_progress_bar=True)
    processed_images = util.paraphrase_mining_embeddings(encoded_image)
    NUM_SIMILAR_IMAGES = 10

    # if(list_of_images[image_id1] == image_1 or list_of_images[image_id2] == image_1):

    duplicates = [image for image in processed_images if image[0] >= 0.99999999]

    for score, image_id1, image_id2 in duplicates[0:NUM_SIMILAR_IMAGES]:
        if (score * 100 == 100.0):
            print("\nScore: {:.3f}%".format(score * 100))
            if (score * 100 > 79.0):
                if (list_of_images[image_id1] == image_1 or list_of_images[image_id2] == image_1):
                    print("Match found")
                    #display(image_1, image_2)
                    path = "static/crop/"
                    shutil.rmtree(path, ignore_errors=True)
                    percentage = score * 100
                    return percentage

    threshold = 0.9999999
    near_duplicates = [image for image in processed_images if image[0] < threshold]
    for score, image_id1, image_id2 in near_duplicates[0:NUM_SIMILAR_IMAGES]:
        print("\nScore: {:.3f}%".format(score * 100))
        if (score * 100 > 79.0):
            if (list_of_images[image_id1] == image_1 or list_of_images[image_id2] == image_1):
                print("Match found")
                #display(image_1, image_2)
                path = "static/crop/"
                shutil.rmtree(path, ignore_errors=True)
                percentage = score * 100
                return percentage

    print("No match!")
    path = "static/crop/"
    shutil.rmtree(path, ignore_errors=True)
    percentage = score * 100
    return percentage

if __name__ == "__main__":
    godEye.run()
