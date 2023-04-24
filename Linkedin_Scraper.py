from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment
from time import sleep
import re
import os.path
import openpyxl
import datetime

def initialisedWebScraping():
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-notifications')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=chrome_options, executable_path=r'chromedriver.exe')
    return driver

def GoogleSearch_LinkedIn(fname,lname,education):
    ## Filter for google dorking
    filters = re.compile(".+linkedin.com/in")
    filteredLink = []

    searchName = fname + " " + lname
    dorking = f'site:linkedin.com AND {searchName}'
    if education:
        dorking = f'site:linkedin.com AND {searchName} AND {education}'

    ## Start Selenium
    driver = initialisedWebScraping()

    ## Navigate to the page specified below
    driver.get('https://www.google.com')

    enterSearchQuery = driver.find_element(By.NAME,'q')
    enterSearchQuery.send_keys(dorking)
    sleep(1)
    enterSearchQuery.send_keys(Keys.RETURN)
    sleep(3)
    count = 1

    try:
        driver.find_element(By.XPATH, '//div[@id="topstuff"]/div/div/p').text
        return "Error"
    except:
        try:
            allLinks = driver.find_elements(By.TAG_NAME,"a")
            for links in allLinks:
                unfilteredLink = links.get_attribute('href')
                if unfilteredLink != None and re.match(filters, '/'.join(unfilteredLink.split('/')[2:4])) and 'trk=public_profile_browsemap' not in unfilteredLink:
                    filteredLink.append(unfilteredLink)

            ## Parse the link login to LinkedInLogin function for scraping later
            return(LinkedInLogin(driver,filteredLink,searchName))
        except:
            return "Error"
def GoogleNextPage(driver):
    nextPageButton = driver.find_element(By.ID,'pnnext')
    nextPageButton.click()

def LinkedInLogin(driver,filteredLink,searchName):
    ## Navigate to the page specified below
    driver.get('https://www.linkedin.com')
    driver.implicitly_wait(3)

    ## To simulate key strokes and enter the email details
    enterEmail = driver.find_element(By.ID,'session_key')
    enterEmail.send_keys('@gmail.com')
    sleep(5)

    ## To simulate key strokes and enter the password
    enterPassword = driver.find_element(By.ID,'session_password')
    enterPassword.send_keys('test')
    sleep(5)

    ## Locate submit button using XPath
    log_in_button = driver.find_element(By.XPATH,'//*[@type="submit"]')
    log_in_button.click()

    return(LinkedInScraping(driver,filteredLink,searchName))

def LinkedInScraping(driver, filteredLink,searchName):
    scrapedInfo = {}
    for profileLink in filteredLink:
        webLink = []
        profileInfo = []
        experiences = []
        educations = []

        ## Redirect to the target's profile
        driver.get(profileLink)
        sleep(1)

        ## Retrieve Profile Image
        profilePic = driver.find_element(By.XPATH,'//div[@class="pv-top-card__non-self-photo-wrapper ml0"]/button/img')
        profilePicLink = profilePic.get_attribute('src')
        if (profilePicLink.startswith('data')):
            profilePicLink = "None"

        ## Retrieve Profile Headline
        headline = driver.find_element(By.XPATH,'//div[@class="text-body-medium break-words"]').text

        ## Retrieve Profile Country
        country = driver.find_element(By.XPATH,'//span[@class="text-body-small inline t-black--light break-words"]').text

        ## Retrieve Experience
        try:
            allExperience = driver.find_elements(By.XPATH,'//*[@id="experience"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li')
            countExp = 1

            for experience in allExperience:
                try:
                    role = experience.find_element(By.XPATH,'//*[@id="experience"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div[1]/div[1]/div/span/span[1]'.format(countExp)).text + "\n"
                except:
                    role = "Unknown" + "\n"
                try:
                    company = experience.find_element(By.XPATH,'//*[@id="experience"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div[1]/div[1]/span[1]/span[1]'.format(countExp)).text + "\n"
                except:
                    company = "Unknown" + "\n"
                try:
                    duration = experience.find_element(By.XPATH,'//*[@id="experience"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div[1]/div[1]/span[2]/span[1]'.format(countExp)).text + "\n"
                except:
                    duration = "Unknown" + "\n"
                try:
                    country = experience.find_element(By.XPATH,'//*[@id="experience"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div[1]/div[1]/span[3]/span[1]'.format(countExp)).text + "\n"
                except:
                    country = "Unknown" + "\n"

                experiences.append(role)
                experiences.append(company)
                experiences.append(duration)
                experiences.append(country)

                countExp += 1
        except:
            experiences.append("None")
            educations.append("None")
            educations.append("None")

        ## Retrieve Education
        try:
            allEducation = driver.find_elements(By.XPATH,'//*[@id="education"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li')
            countEdu = 1

            for education in allEducation:
                try:
                    school = education.find_element(By.XPATH,'//*[@id="education"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div/a/div/span/span[1]'.format(countEdu)).text + "\n"
                except:
                    school = "Unknown" + "\n"
                try:
                    qualification = education.find_element(By.XPATH,'//*[@id="education"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div/a/span[1]/span[1]'.format(countEdu)).text + "\n"
                except:
                    qualification = "Unknown" + "\n"
                try:
                    duration = education.find_element(By.XPATH,'//*[@id="education"]//following-sibling::div[@class="pvs-list__outer-container"]/ul/li[{}]/div/div[2]/div/a/span[2]/span[1]'.format(countEdu)).text + "\n"
                except:
                    duration = "Unknown" + "\n"

                educations.append(school)
                educations.append(qualification)
                educations.append(duration)

                countEdu += 1
        except:
            educations.append("None")
            educations.append("None")
            educations.append("None")

        ## Retrieve Additional Contact Info if any
        additionalContactInfo = driver.find_element(By.ID,'top-card-text-details-contact-info')
        additionalContactInfo.click()
        sleep(1)

        ## Retrieve the profile's name
        name = driver.find_element(By.ID,'pv-contact-info').text

        ## Add information into a list and to be added into a dictionary later
        profileInfo.append(name)
        profileInfo.append(profilePicLink)
        profileInfo.append(headline)
        profileInfo.append(country)
        profileInfo.append(experiences)
        profileInfo.append(educations)

        ## Retrieve the 'Website' Link
        try:
            additionalWebLink = driver.find_elements(By.XPATH,'//section[@class="pv-contact-info__contact-type ci-websites"]/ul/li[@class="pv-contact-info__ci-container link t-14"]/a')
            for link in additionalWebLink:
                webLink.append(link.get_attribute('href')+"\n")
        except:
            webLink.append('None')
        profileInfo.append(webLink)

        ## Retrieve the Phone number (If Applicable)
        try:
            phoneNumber = driver.find_element(By.XPATH,'//section[@class="pv-contact-info__contact-type ci-phone"]/ul/li[@class="pv-contact-info__ci-container link t-14"]/span[@class="t-14 t-black t-normal"]').text
        except:
            phoneNumber = "None"
        profileInfo.append(phoneNumber)

        ## Retrieve the Address (If Applicable)
        try:
            address = driver.find_element(By.XPATH,'//section[@class="pv-contact-info__contact-type ci-address"]/div[@class="pv-contact-info__ci-container t-14"]/a').text
        except:
            address = "None"
        profileInfo.append(address)

        ## Retrieve the Email Address
        try:
            unfilteredEmail = driver.find_element(By.XPATH,'//section[@class="pv-contact-info__contact-type ci-email"]/div[@class="pv-contact-info__ci-container t-14"]/a')
            email = unfilteredEmail.get_attribute('href')
        except:
            email = "None"
        profileInfo.append(email)

        ## Update the profile information into a dictionary
        scrapedInfo.update({profileLink:profileInfo})
        sleep(1)

    driver.get('https://linkedin.com/m/logout')
    driver.close()
    return(writeToExcel(scrapedInfo,searchName))

def writeToExcel(scrapedInfo,searchName):
    filename = str(datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")).replace("-","") + " " + searchName + " LinkedinData.xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        workbook.save(filename)

    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.worksheets[0]
    header = ["Profile Link","Name","Profile Pic Link","Headline","Country","Role","Company","Duration (Work)","Country","Education","Qualification","Duration (School)","Additional Links","Phone Number","Address","Email"]

    column = 1

    ## Write Header content to excel
    for headerData in header:
        worksheet.cell(1,column).value = headerData
        column += 1
    row = 2

    ## Write the dictionary contents into the excel
    for key,allData in scrapedInfo.items():
        column = 1
        worksheet.cell(row,column).value = key
        column += 1
        workCount = 0
        eduCount = 0

        for data in allData:
            if column == 6:
                while data:
                    workCount += 1
                    try:
                        value = worksheet.cell(row,column).value
                        newValue = value + data[0]
                    except:
                        newValue = data[0]
                    worksheet.cell(row,column).value = newValue
                    if workCount == 4:
                        column = 5
                        workCount = 0
                    column+=1
                    data.pop(0)
                if data == []:
                    worksheet.cell(row,6).alignment = Alignment(wrapText=True)
                    worksheet.cell(row,7).alignment = Alignment(wrapText=True)
                    worksheet.cell(row,8).alignment = Alignment(wrapText=True)
                    worksheet.cell(row,9).alignment = Alignment(wrapText=True)
                    column = 9

            elif column == 10:
                while data:
                    eduCount += 1
                    try:
                        value = worksheet.cell(row,column).value
                        newValue = value + data[0]
                    except:
                        newValue = data[0]
                    worksheet.cell(row,column).value = newValue
                    if eduCount == 3:
                        column = 9
                        eduCount = 0
                    column+=1
                    data.pop(0)
                if data == []:
                    worksheet.cell(row,10).alignment = Alignment(wrapText=True)
                    worksheet.cell(row,11).alignment = Alignment(wrapText=True)
                    worksheet.cell(row,12).alignment = Alignment(wrapText=True)
                    column = 12

            elif column == 13:
                worksheet.cell(row,column).value =''.join(data)

            else:
                worksheet.cell(row,column).value = data
            column += 1
        row += 1

    workbook.save(filename)
    workbook.close()
    return filename
